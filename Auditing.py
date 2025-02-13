import tempfile
import pandas as pd
import os
import re

import zipfile

from lxml import etree

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.table import Table, TableStyleInfo

import mmap
from concurrent.futures import ProcessPoolExecutor

from io import BytesIO
import requests
from requests.auth import HTTPBasicAuth
from requests.packages.urllib3.exceptions import InsecureRequestWarning # type: ignore

import warnings
warnings.simplefilter('ignore', InsecureRequestWarning)

"""
Ideas to look into:
    * Creating a web interface for the user to interact with.
        - This would require possible api work to get lists of site and asset folders for dropdowns.
        - This could fix the issue of having to run the script everytime if we can just put this on the server.
        - Using a python framework called Flask the web interface can be created.
    * Find somewhere to put the code so someone else can use it remotely.
        - Like mentioned above if it can get added to the server somehow that would be ideal.
        - If it can't be added onto python though the code has to get transferred to coldfusion and added like that.
"""

"""
Instructions for setting up an audit:

    1.) In This Code:
            - Change 4 variables below according to what site you're auditing.

            - PackageName: Should be name of site followed by Audit. Like 'GreenhouseAudit'.
            - WebpageURL: Is the Web url that comes before pages. Like 'greenhouses.caes.uga.edu'.
            - AEM_Sites_Folder_Path: This is the place/path of the site you want to audit on AEM. Like '/content/caes-subsite/greenhouses'.
            - AEM_Assets_Folder_Path: This is the place/path where the assets of your site live on AEM. Like '/content/dam/caes-subsite/greenhouses'.

    2.) Running could take a couple minutes. After the code runs check to make sure workbook looks correct and send off!

"""
#############

"""
PackageName = 'GreenhouseAudit'
WebpageURL = 'greenhouses.caes.uga.edu'
AEM_Sites_Folder_Path = '/content/caes-subsite/greenhouses'
AEM_Assets_Folder_Path = '/content/dam/caes-subsite/greenhouses'
print("\nThank You!\n")
"""
#PackageName = 'ExtCountyOfficesAudit'
#WebpageURL = 'extension.uga.edu/county-offices' # Needs to be exact url!
#AEM_Sites_Folder_Path = '/content/extension/county-offices' # Needs to be exact path!
#AEM_Assets_Folder_Path = '/content/dam/extension-county-offices' # Needs to be exact path!


###############################################
def get_asset_info(asset_path, aem_host, username, password):
    asset_info = {}
    metadata_url = f'{aem_host}{asset_path}/jcr:content/metadata.json'
    metadata_response = requests.get(metadata_url, auth=HTTPBasicAuth(username, password), verify=False)

    replication_url = f'{aem_host}{asset_path}/jcr:content.json'
    replication_response = requests.get(replication_url, auth=HTTPBasicAuth(username, password), verify=False)

    if metadata_response.status_code == 200:
        metadata = metadata_response.json()

        if replication_response.status_code == 200:
            replication_data = replication_response.json()
            if replication_data.get('cq:lastReplicationAction') == "Activate":
                asset_info['Published?'] = 'Yes'
            elif replication_data.get('cq:lastReplicationAction') == None:
                asset_info['Published?'] = 'No'
            else:
                asset_info['Published?'] = 'ERROR'
        asset_info['Path'] = asset_path
        asset_info['File Size (in Bytes)'] = metadata.get('dam:size', '0')

    return asset_info
###############################################

###############################################
def get_assets_through_api(AEM_Assets_Folder_Path, aem_host, username, password):
    assets = []
    query_url = (f'{aem_host }/bin/querybuilder.json?path={AEM_Assets_Folder_Path}&type=dam:Asset&p.limit=-1')
    response = (requests.get(query_url, auth=HTTPBasicAuth(username, password), verify=False))

    if response.status_code == 200:
        results = response.json().get('hits', [])
        for result in results:
            asset_path = result.get('path')
            asset_info = get_asset_info(asset_path, aem_host, username, password)
            if asset_info:
                assets.append(asset_info)
    return assets
###############################################


###############################################
def getContentFilePaths(main_content_folder):
    content_files = []
    folder_paths = []

    for root, dirs, files in os.walk(main_content_folder):

        dirs[:] = [d for d in dirs if d != '_jcr_content']


        short_root = os.path.relpath(root)

        for file in files:
            if file.endswith('.content.xml') and not any(substring in short_root for substring in ['\\search', '\\sitemap', '\\news', '\\calendar']):
                folder_paths.append(short_root)
                content_files.append(os.path.join(short_root, file))

    return content_files, folder_paths
###############################################

###############################################
def getWebpageInfo(list_of_files, folder_paths, parent_folder_path):

    #WebpageURL = user_webpage_info()

    for i in range(len(folder_paths)):
        folder_paths[i] = folder_paths[i].replace('\\', '/')

    pages = []
    url_array = []

    namespace_mapping = {
            'jcr': 'http://www.jcp.org/jcr/1.0',
            'cq': 'http://www.day.com/jcr/cq/1.0'
        }

    for file in list_of_files:
        tree = etree.parse(file)
        root = tree.getroot()

        matching_elements = root.xpath('//jcr:content', namespaces=namespace_mapping)

        for contentElement in matching_elements:
            if contentElement.get('{http://www.jcp.org/jcr/1.0}title'):
                pageInfo = {}
                if contentElement.get('{http://www.day.com/jcr/cq/1.0}lastReplicationAction') == 'Activate':
                    pageInfo['Published?'] = 'Yes'
                else: pageInfo['Published?'] = 'No'

                pageInfo['Title'] = contentElement.get('{http://www.jcp.org/jcr/1.0}title')
                pageInfo['Description'] = contentElement.get('{http://www.jcp.org/jcr/1.0}description')
                if pageInfo['Description'] == None:
                    pageInfo['Description'] = ' '

                pages.append(pageInfo)

    webpage_df = pd.DataFrame(pages)

    for i in folder_paths:
        ########## Change URL to correct site URL ##########
        i = i.replace(parent_folder_path, '')
        url_array.append(f'{WebpageURL}{i[i.find(os.path.basename(parent_folder_path)) + len(os.path.basename(parent_folder_path)):]}' )
        #print(i)
    if len(webpage_df) == 0:
        print('webpage_df empty...')
        exit(1)
    else:
        webpage_df.insert(1, 'URLs', url_array)

    print("WebPage Dataframe: ")
    print(webpage_df)
    print()
    print()

    return webpage_df
###############################################

###############################################
def search_file_content(file_path, target_files, namespace_mapping):
    file_info = []
    with open(file_path, 'r', encoding='utf-8') as file:
        # Use memory-mapped file for efficient search
        with mmap.mmap(file.fileno(), 0, access=mmap.ACCESS_READ) as m:
            content = m.read().decode('utf-8')
            for target_file in target_files:
                space_target_file = target_file.replace(' ', '%20')
                if (target_file in content or space_target_file in content):
                    if target_file is not None and target_file != '':
                        #print(target_file)
                        info = {'fileReference': target_file}
                        match = re.search(r'jcr:title="([^"]+)"', content)
                        if match:
                            title_value = match.group(1)
                            content = re.sub(r'^<\?xml.*?\?>', '', content).strip()
                            element = etree.XML(content.encode())
                            etree.register_namespace('jcr', namespace_mapping['jcr'])
                            attribute = element.xpath('//@jcr:title', namespaces=namespace_mapping)
                            page_title = attribute[0] if attribute else title_value
                            info['referencedBy'] = page_title
                        file_info.append(info)
    return file_info
###############################################

###############################################
def search_files_for_string(directory, FilesFromCSV):
    found_files = []

    namespace_mapping = {'jcr': 'http://www.jcp.org/jcr/1.0',}

    # Builds a list of all .content.xml files once to avoid building again.
    all_files = []
    for root, dirs, files in os.walk(directory):
        dirs[:] = [d for d in dirs if d != '_jcr_content']
        short_root = os.path.relpath(root)
        for file in files:
            if file.endswith('.content.xml'):
                file_path = os.path.join(short_root, file)
                all_files.append(file_path)

    # ThreadPool/ProcessPoolExecutor helps run things simultaniously.
    # Processes files in parallel batches
    batch_size = 100
    with ProcessPoolExecutor() as executor:
        for i in range(0, len(all_files), batch_size):
            batch_files = all_files[i:i + batch_size]
            future_to_file = {executor.submit(search_file_content, file_path, FilesFromCSV, namespace_mapping): file_path for file_path in batch_files}
            for future in future_to_file:
                result = future.result()
                if result:
                    found_files.extend(result)

    return found_files
###############################################


###############################################
def autosize_ws_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if column_letter == 'H':
                break
            else:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
        adjusted_width = (max_length) + 3
        ws.column_dimensions[column_letter].width = adjusted_width
###############################################


###############################################
def main():

    #PackageName, AEM_Sites_Folder_Path, AEM_Assets_Folder_Path = user_AEM_info()

    ####### Getting parent_folder #######

    # Main URL
    # https://aem.caes.uga.edu:8443/crx/packmgr/service.jsp

    # Start of URL
    # https://aem.caes.uga.edu:8443/crx/packmgr/service/.json

    # Create
    # https://aem.caes.uga.edu:8443/crx/packmgr/service/.json?cmd=create&groupName=site_audits&packageName=GreenhouseAudit2&version=1.0

    # Build
    # https://aem.caes.uga.edu:8443/crx/packmgr/service/.json/etc/packages/site_audits/GreenhouseAudit2.zip?cmd=build

    # Download
    # https://aem.caes.uga.edu:8443/etc/packages/site_audits/GreenhouseAudit2.zip

    # Upload
    # https://aem.caes.uga.edu:8443/crx/packmgr/service/.json?cmd=upload

    # Delete
    # https://aem.caes.uga.edu:8443/crx/packmgr/service.jsp?cmd=rm&name=GreenhouseAudit2


    aem_host = 'https://aem.caes.uga.edu:8443'
    username = 'admin'
    password = 'Mt4sbwu'


    group_name = 'site_audits'
    version = '1.0'

    package_manager_url = f'{aem_host}/crx/packmgr/service/.json'
    package_file_path = f'/etc/packages/{group_name}/{PackageName}.zip'


    #### Creating initial package ####
    create_package_url = f'{package_manager_url}'
    data = {
        'cmd': 'create',
        'groupName': group_name,
        'packageName': PackageName,
        'version': version
    }

    response = requests.post(create_package_url, data=data, auth=HTTPBasicAuth(username, password), verify=False)

    if response.status_code == 200:
        response_json = response.json()
        if response_json.get('success'):
            print(f'Package {PackageName} created successfully.')
        elif response_json.get("msg") == f'{PackageName}.zip':
            print(f'Package already exists...using that data now')
        else:
            print(f'Failed to create package: {response_json.get("msg")}')
    else:
        print(f'Failed to create package. Status code: {response.status_code}')


    #### Building the empty package we just created ####
    ## This makes sure all the neccessary files are built and can be uploaded later ##
    build_package_url = f'{package_manager_url}{package_file_path}?cmd=build'

    response = requests.post(build_package_url, auth=HTTPBasicAuth(username, password), verify=False)

    if response.status_code == 200:
        print(f'Package {PackageName} built successfully.')
    else:
        print(f'Failed to build package. Status code: {response.status_code}')


    #### Downloading empty package to add to it's filter.xml ####
    download_package_url = f'{aem_host}{package_file_path}'

    response = requests.get(download_package_url, auth=(username, password), verify=False)

    if response.status_code == 200:
        package_content = BytesIO(response.content)
        print(f'Package {PackageName} downloaded using BytesIO successfully.')
    else:
        print(f'Failed to download package. Status code: {response.status_code}')
        print(f'Response content: {response.text}')


    #### Remove the package off AEM to keep AEM clean ####
    remove_package_url = f'{aem_host}/crx/packmgr/service.jsp?cmd=rm&name={PackageName}'

    response = requests.post(remove_package_url, auth=HTTPBasicAuth(username, password), verify=False)

    if response.status_code == 200:
        print(f'Package {PackageName} removed successfully.')
    else:
        print(f'Failed to remove package. Status code: {response.status_code}')


    #### Use BytesIO to manipulate and add correct filter to filter.xml ####
    filter_root = [{'root': AEM_Sites_Folder_Path, 'excludes': [(AEM_Sites_Folder_Path+'/404'), (AEM_Sites_Folder_Path+'/403')]}]
    filter_entries = (''.join([f'<filter root="{f["root"]}">\n\n' + ''.join([f'\t\t<exclude pattern="{e}"/>\n' for e in f["excludes"]]) + '\n\t</filter>\n' for f in filter_root]))
    filter_xml_content = f'<?xml version="1.0" encoding="UTF-8"?>\n<workspaceFilter version="1.0">\n\t{filter_entries}</workspaceFilter>'

    new_package_data = BytesIO()

    with zipfile.ZipFile(package_content, 'r') as original_zip:
        with zipfile.ZipFile(new_package_data, 'w') as new_zip:
            for item in original_zip.infolist():
                if item.filename != 'META-INF/vault/filter.xml':
                    new_zip.writestr(item, original_zip.read(item.filename))

            new_zip.writestr('META-INF/vault/filter.xml', filter_xml_content.encode('utf-8'))

    new_package_data.seek(0)


    #### Upload the newpackage_data zip to AEM ####
    upload_package_url = f'{aem_host}/crx/packmgr/service/.json?cmd=upload'

    files = {'package': ('new.zip', new_package_data, 'application/zip')}

    response = requests.post(upload_package_url, files=files, auth=HTTPBasicAuth(username, password), verify=False)

    if response.status_code == 200 and 'success' in response.json() and response.json()['success']:
        print('Package uploaded and installed successfully.')
    else:
        print(f'Failed to upload package. Status code: {response.status_code}')
        print(f'Response content: {response.text}')
    print()


    #### Building the new and improved package ####
    build_package_url = f'{package_manager_url}{package_file_path}?cmd=build'

    response = requests.post(build_package_url, auth=HTTPBasicAuth(username, password), verify=False)

    if response.status_code == 200:
        print(f'Package {PackageName} built successfully.')
    else:
        print(f'Failed to build package. Status code: {response.status_code}')


    #### Download and use ByteIO() as the parent_folder_path ####
    download_package_url = f'{aem_host}{package_file_path}'

    response = requests.get(download_package_url, auth=(username, password), verify=False)

    if response.status_code == 200:
        final_package_content = BytesIO(response.content)
        print(f'Final package {PackageName} downloaded using BytesIO successfully.')
    else:
        print(f'Failed to download package. Status code: {response.status_code}')
        print(f'Response content: {response.text}')

    #### Creating temp directory to unzip package. ####
    with tempfile.TemporaryDirectory() as temp_dir:
        print(f'Creating temporary directory called {temp_dir}...')
        print()
        with zipfile.ZipFile(final_package_content, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        new_AEM_Sites_Folder_Path = AEM_Sites_Folder_Path.replace('/', '\\')
        parent_folder_path = os.path.join(temp_dir + '\\jcr_root' + new_AEM_Sites_Folder_Path)

    #####################################
    #### Start of setting up actual information ####
        audit_folder = 'Audits'
        title_word = PackageName.split('Audit')[0]
        if not os.path.exists(audit_folder):
            print('Making workbook folder...')
            os.mkdir(audit_folder)
        WorkbookName = (rf'{audit_folder}\{title_word}_Website_Audit_Workbook.xlsx')


        ##### Getting the Assets #####
        print('Getting assets from AEM...')
        new_assets_df = pd.DataFrame((get_assets_through_api(AEM_Assets_Folder_Path, aem_host, username, password)))
        #print(new_assets_df)
        FilesFromCSV = (new_assets_df['Path'])
        #for asset in assets:
            #print(f"Path: {asset['path']}, Published: {asset['published']}, Size: {asset['size']}")
        print('These are the assets gotten through the API:')
        print(new_assets_df)
        ##############################


        ######## Calling Functions ########
        print('Getting lists of .content.xml files and folder paths...')
        content_files, folder_paths = getContentFilePaths(parent_folder_path)

        print('Getting the page and subpage info...')
        pages_df = getWebpageInfo(content_files, folder_paths, parent_folder_path)

        print('Searching through webpage parent folder for file references...')
        ReferencedFiles = search_files_for_string(parent_folder_path, FilesFromCSV)
        ref_df = pd.DataFrame(ReferencedFiles)

        asset_df_subset = (new_assets_df)
        asset_sheet_df = pd.concat([asset_df_subset, ref_df], axis=1)

        print('Building Workbook...')
        asset_sheet_df.insert(1, 'Name', pd.Series(dtype=object))
        asset_sheet_df.insert(3, 'Referenced By', pd.Series(dtype=object))
        asset_sheet_df.insert(4, 'File Extension', pd.Series(dtype=object))
        asset_sheet_df.insert(5, 'File Size (in MB)', pd.Series(dtype=object))
        asset_sheet_df.insert(7, ' ', pd.Series(dtype=object))

        print('Actual used fileReference Dataframe: ')
        print(ref_df)
        print()
        print('Final Asset Dataframe uploaded to workbook without some info: ')
        print(asset_sheet_df)
        print()

        print("\nDone!")
        print("The workbook should be in the 'Audits' folder!")
        #asset_sheet_df.to_csv('new_csv.csv', index=False)


        with pd.ExcelWriter(WorkbookName, engine='openpyxl') as writer:
            pages_df.to_excel(writer, sheet_name='Webpages', index=False)
            asset_sheet_df.to_excel(writer, sheet_name='Assets', index=False)

        wb = load_workbook(WorkbookName)
        ws = wb['Webpages']

        ws['E1'] = "Action"
        ws['F1'] = "Notes to Modify"


        ##### Resets first column formatting #####
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font()
            cell.border = openpyxl.styles.Border()
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            cell.alignment = openpyxl.styles.Alignment()
            cell.number_format = 'General'
        ##########################################


        ##### Webpage Rules #####
        fillNos = PatternFill(start_color='FF8989', end_color='FF8989', fill_type='solid')
        noRule = FormulaRule(formula=['$A2="No"'], stopIfTrue=False, fill=fillNos)

        ws.conditional_formatting.add(f'A2:F{ws.max_row}', noRule)
        #########################


        ##### Making Webpage table #####
        table_range = f'A1:F{ws.max_row}'
        webpageTable = Table(displayName='WebpagesTable', ref=table_range)
        style = TableStyleInfo(name='TableStyleLight15', showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)

        webpageTable.tableStyleInfo = style
        ws.add_table(webpageTable)
        ################################

        autosize_ws_columns(ws)

        ##### Filling Asset sheet missing info #####
        ws = wb['Assets']
        for row in range(2, ws.max_row + 1):
            if ws[f'C{row}'].value:
                file_path = ws[f'C{row}'].value
                ws[f'B{row}'] = os.path.basename(file_path)
                ws[f'E{row}'] = os.path.splitext(file_path)[1]

                for cell in range(2, ws.max_row + 1):
                    if ws[f'I{cell}'].value is not None and ws[f'I{cell}'].value != 'fileReference':
                        file_ref = ws[f'I{cell}'].value
                        if file_ref == file_path:
                            if ws[f'D{row}'].value is None:
                                ws[f'D{row}'] = ws[f'J{cell}'].value
                            else:
                                ws[f'D{row}'] = f"{ws[f'D{row}'].value}, {ws[f'J{cell}'].value}"


            if ws[f'G{row}'].value is not None:
                file_bytes = int(ws[f'G{row}'].value)
                ws[f'F{row}'] = float(f'{file_bytes/(1024*1024):.3f}')
                ws[f'G{row}'] = int(ws[f'G{row}'].value)
        ################################

        autosize_ws_columns(ws)
        ws.column_dimensions.group(start='H', end='J', hidden=True)


        ##### Resets first column formatting #####
        for cell in ws[1]:
            cell.font = openpyxl.styles.Font()
            cell.border = openpyxl.styles.Border()
            cell.fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
            cell.alignment = openpyxl.styles.Alignment()
            cell.number_format = 'General'
        ##########################################


        ###### Asset Rules ######
        fill_images = PatternFill(start_color='FFF3AB', end_color='FFF3AB', fill_type='solid')
        fill_pdfs = PatternFill(start_color='F8D0BA', end_color='F8D0BA', fill_type='solid')

        imageRule = FormulaRule(formula=['AND(OR($E2=".png", $E2=".jpg", $E2=".jpeg"), $F2>0.5)'], stopIfTrue=False, fill=fill_images)
        pdfRule = FormulaRule(formula=['AND(OR($E2=".pdf", $E2=".docx"), $F2>1)'], stopIfTrue=False, fill=fill_pdfs)
        assetNoRule = FormulaRule(formula=['OR($D2="", $A2="No")'], stopIfTrue=False, fill=fillNos)

        ws.conditional_formatting.add(f'A2:G{ws.max_row}', assetNoRule)
        ws.conditional_formatting.add(f'A2:G{ws.max_row}', imageRule)
        ws.conditional_formatting.add(f'A2:G{ws.max_row}', pdfRule)
        #########################


        ###### Making Asset Table ######
        table_range = f'A1:G{ws.max_row}'
        webpageTable = Table(displayName='AssetsTable', ref=table_range)
        style = TableStyleInfo(name='TableStyleLight15', showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)

        webpageTable.tableStyleInfo = style
        ws.add_table(webpageTable)
        ################################

        wb.create_sheet('Color Key')
        ws = wb['Color Key']
        ws['B2'] = "On a Webpage, red means the page isn't published."
        ws['B2'].fill = fillNos
        ws['B3'] = "On an Asset, red means the asset isn't published or isn't referenced."
        ws['B3'].fill = fillNos
        ws['B4'] = "On an Asset, yellow means the image is over 500 KB."
        ws['B4'].fill = fill_images
        ws['B5'] = "On an Asset, orange means the pdf is over 1 MB."
        ws['B5'].fill = fill_pdfs
        for cell in range(2, ws.max_row + 1):
            ws[f'B{cell}'].border = Border(left=Side('medium'), right=Side('medium'), top=Side('medium'), bottom=Side('medium'))

        autosize_ws_columns(ws)

        wb.save(WorkbookName)
        ######### End of calling functions #########
###############################################

if __name__ == '__main__':
    """ PackageName = input("Package Name: ")
    WebpageURL = 'greenhouses.caes.uga.edu'
    AEM_Sites_Folder_Path = input("AEM Site Folder Path: ")
    AEM_Assets_Folder_Path = input("AEM Asset Folder Path: ")
    print("\nThank You!\n") """

    PackageName = 'GreenhouseAudit'
    WebpageURL = 'greenhouses.caes.uga.edu'
    AEM_Sites_Folder_Path = '/content/caes-subsite/greenhouses'
    AEM_Assets_Folder_Path = '/content/dam/caes-subsite/greenhouses'

    print("\nThank You!\n")

    main()